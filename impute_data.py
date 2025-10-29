import pandas as pd
import requests
from datetime import datetime
import openpyxl
import toml
import os

def load_config():
    try:
        env_key = os.getenv('TIINGO_API_KEY', '')
        if env_key:
            return {
                'tiingo': {
                    'api_key': env_key
                }
            }
        
        if os.path.exists('secrets.toml'):
            secrets = toml.load('secrets.toml')
            
            if 'tiingo' in secrets and 'api_key' in secrets['tiingo']:
                return secrets
                
            if 'tiingo_api_key' in secrets:
                return {
                    'tiingo': {
                        'api_key': secrets['tiingo_api_key']
                    }
                }
        
        return {
            'tiingo': {
                'api_key': ''
            }
        }
        
    except Exception as e:
        print(f"Error loading configuration: {e}")
        return {
            'tiingo': {
                'api_key': ''
            }
        }

config = load_config()
API_KEY = config.get('tiingo', {}).get('api_key', '')

if not API_KEY:
    raise ValueError("Tiingo API key not found in secrets.toml or environment variables")

headers = {'Content-Type': 'application/json', 'Authorization': f'Token {API_KEY}'}

tickers = ['ACWI', 'AGG', 'SPY']
portfolios = {
    '50/50': {'ACWI': 0.5, 'AGG': 0.5},
    '60/40': {'ACWI': 0.6, 'AGG': 0.4},
    '70/30': {'ACWI': 0.7, 'AGG': 0.3},
}

start = datetime(2019, 7, 1)
today = datetime.today()
dates = pd.date_range(start, today, freq='M')

def get_monthly_adj_returns(ticker):
    url = f'https://api.tiingo.com/tiingo/daily/{ticker}/prices'
    params = {
        'startDate': dates[0].strftime('%Y-%m-%d'),
        'endDate': dates[-1].strftime('%Y-%m-%d'),
        'resampleFreq': 'daily',
        'columns': 'adjClose,date'
    }
    r = requests.get(url, headers=headers, params=params)
    data = r.json()
    df = pd.DataFrame(data)
    df['date'] = pd.to_datetime(df['date']).dt.tz_localize(None)
    df = df.set_index('date').sort_index()
    monthly = df['adjClose'].resample('M').last()
    returns = monthly.pct_change()
    if not returns.empty:
        returns.iloc[0] = 0.0
    return returns

returns_dict = {t: get_monthly_adj_returns(t) for t in tickers}

def get_portfolio_return(weights, returns_dict):
    df = pd.DataFrame({t: returns_dict[t] for t in weights})
    port_ret = (df * pd.Series(weights)).sum(axis=1)
    return port_ret

portfolio_returns = {name: get_portfolio_return(w, returns_dict) for name, w in portfolios.items()}

wb = openpyxl.load_workbook('GA_new copy.xlsx')
ws = wb.active

col_map = {
    'ACWI': 4,   # D
    'AGG': 5,    # E
    'SPY': 6,    # F
    '50/50': 7,  # G
    '60/40': 8,  # H
    '70/30': 9   # I
}

import numpy as np
row_map = {}
for row in range(2, ws.max_row+1):
    date_val = ws.cell(row=row, column=1).value
    if isinstance(date_val, datetime):
        date_str = date_val.strftime('%-m/%-d/%y')
    elif isinstance(date_val, str):
        try:
            dt = datetime.strptime(date_val, '%m/%d/%y')
            date_str = dt.strftime('%-m/%-d/%y')
        except Exception:
            try:
                dt = datetime.strptime(date_val, '%-m/%-d/%y')
                date_str = dt.strftime('%-m/%-d/%y')
            except Exception:
                continue
    else:
        continue
    row_map[date_str] = row

for date in returns_dict['ACWI'].index:
    date_str = date.strftime('%-m/%-d/%y')
    row = row_map.get(date_str)
    if not row:
        continue
    
    for t in tickers:
        val = returns_dict[t].get(date, None)
        if val is not None and not (isinstance(val, float) and (np.isnan(val) or np.isinf(val))):
            ws.cell(row=row, column=col_map[t]).value = round(val, 4)
    
    for p in portfolios:
        val = portfolio_returns[p].get(date, None)
        if val is not None and not (isinstance(val, float) and (np.isnan(val) or np.isinf(val))):
            ws.cell(row=row, column=col_map[p]).value = round(val, 4)

wb.save('GA_new copy.xlsx')
print("Excel file updated successfully!")