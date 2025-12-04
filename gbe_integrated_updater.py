"""
GBE Strategy Integrated Monthly Updater
"""

import pandas as pd
import openpyxl
from datetime import date, datetime
import calendar
import os
import requests
import time
import gzip
import io
import csv
from xml.etree import ElementTree as ET
import toml

class GBEMonthlyUpdater:
    """Integrated GBE monthly data processor"""
    
    def _load_secrets(self):
        """Load secrets from secrets.toml or Streamlit secrets"""
        try:
            import streamlit as st
            if hasattr(st, 'secrets'):
                return dict(st.secrets)
        except:
            pass
        
        if os.path.exists('secrets.toml'):
            return toml.load('secrets.toml')
        
        raise FileNotFoundError("No secrets.toml file found and not running in Streamlit")
    
    def __init__(self):
        # GBE ETF mapping - aggregates multiple ETFs into canonical asset classes
        self.etf_mapping = {
            # Commodities
            "PDBC": "PDBC",
            "DBC": "PDBC",
            "COMT": "PDBC",
            
            # Emerging Markets
            "VWO": "VWO",
            "SCHE": "VWO",
            
            # International Developed (Europe)
            "IEV": "IEV",
            
            # International Bonds
            "BWX": "BWX",
            
            # Gold
            "IAU": "IAU",
            "GLD": "IAU",
            
            # International Developed (VEA/EFA combined)
            "VEA": "VEA",
            "EFA": "VEA",
            
            # International REITs
            "RWX": "RWX",
            
            # Japan Bonds
            "EWJ": "EWJ",
            
            # Intermediate Bonds
            "BIV": "BIV",
            
            # S&P 500
            "SPY": "SPY",
            "SPLG": "SPY",
            "SPYM": "SPY",
            
            # TIPS (Inflation Protected)
            "TIP": "TIP",
            
            # Aggregate Bonds
            "BND": "BND",
            "AGG": "BND",
            "IAGG": "BND",  # International aggregate bonds -> BND
            
            # Long Term Treasuries
            "SPTL": "SPTL",
            "TLT": "SPTL",
            
            # Intermediate Treasuries (7-10 year)
            "IEF": "BIV",  # 7-10 year treasuries -> Intermediate Bonds
            
            # TIPS (Inflation Protected)
            "SCHP": "TIP",  # Schwab TIPS ETF -> TIP
            
            # REITs
            "VNQ": "VNQ",
            "IYR": "VNQ",
        }
        
        # Canonical ETFs for GBE (order matches spreadsheet columns)
        # 14 ETFs matching the actual positions in the portfolio
        self.canonical_etfs = ["PDBC", "VWO", "IEV", "BWX", "IAU", "VEA", "RWX", "EWJ", "BIV", "SPY", "TIP", "BND", "SPTL", "VNQ"]
        
        # IBKR API Configuration - Load from secrets
        secrets = self._load_secrets()
        self.flex_token = secrets.get('ibkr_gbe', {}).get('flex_token', '')
        self.query_id = secrets.get('ibkr_gbe', {}).get('query_id', '')
        self.primary_account = secrets.get('ibkr_gbe', {}).get('primary_account', '')
        
        if not all([self.flex_token, self.query_id, self.primary_account]):
            raise ValueError("IBKR GBE credentials not found in secrets.toml")
        
        self.send_url = "https://www.interactivebrokers.com/Universal/servlet/FlexStatementService.SendRequest"
        self.get_url = "https://www.interactivebrokers.com/Universal/servlet/FlexStatementService.GetStatement" 
    
    def fetch_flex_data_direct(self):
        """Fetch flex data directly from IBKR API without file intermediates"""
        
        print("Fetching data directly from IBKR Flex Query API...")
        print(f"Token: {self.flex_token}")
        print(f"Query ID: {self.query_id}")
        
        session = requests.Session()
        headers = {"User-Agent": "GBE-FlexPipeline/1.0"}
        
        try:
            print("Step 1: Sending flex query request...")
            params = {"t": self.flex_token, "q": self.query_id, "v": "3"}
            response = session.get(self.send_url, params=params, timeout=30, headers=headers)
            
            if response.status_code != 200:
                raise Exception(f"SendRequest failed with HTTP {response.status_code}: {response.text[:500]}")
            
            root = ET.fromstring(response.text)
            status = root.findtext(".//Status") or root.findtext(".//status") or ""
            
            if status.lower() == "fail":
                error_code = root.findtext(".//ErrorCode") or ""
                error_msg = root.findtext(".//ErrorMessage") or ""
                raise Exception(f"Flex request failed: {error_code} - {error_msg}")
            
            ref_code = root.findtext(".//ReferenceCode") or root.findtext(".//referenceCode")
            if not ref_code:
                raise Exception(f"No reference code in response: {response.text[:500]}")
            
            print(f"✅ Got reference code: {ref_code}")
            
            print("Step 2: Polling for results...")
            max_attempts = 10
            
            for attempt in range(1, max_attempts + 1):
                time.sleep(3 if attempt > 1 else 2)
                
                params = {"t": self.flex_token, "q": ref_code, "v": "3"}
                response = session.get(self.get_url, params=params, timeout=60, headers=headers)
                
                if response.status_code != 200:
                    print(f"Attempt {attempt}: HTTP {response.status_code}")
                    continue
                
                if response.text.lstrip().startswith("<"):
                    try:
                        root = ET.fromstring(response.text)
                        status = root.findtext(".//Status") or root.findtext(".//status") or ""
                        
                        if status.lower() in {"success", "ready"}:
                            content = response.content
                        else:
                            print(f"Attempt {attempt}: Status = {status}")
                            continue
                    except ET.ParseError:
                        content = response.content
                else:
                    content = response.content
                
                if response.headers.get("Content-Encoding", "").lower() == "gzip" or content[:2] == b"\x1f\x8b":
                    try:
                        content = gzip.decompress(content)
                    except Exception:
                        pass
                
                print(f"✅ Retrieved data ({len(content)} bytes)")
                return content.decode("utf-8", errors="replace")
            
            raise Exception(f"Data not ready after {max_attempts} attempts")
            
        except Exception as e:
            print(f"❌ API Error: {e}")
            raise
        
    def extract_all_weather_data_direct(self):

        print("GBE Integrated Monthly Updater")
        print("=" * 50)
        print("Extracting ALL AllWeather accounts...")
        
        content = self.fetch_flex_data_direct()
        
        # Debug: Save raw content to file for inspection
        debug_file = f"debug_flex_data_gbe_{date.today().strftime('%Y%m%d')}.csv"
        with open(debug_file, 'w') as f:
            f.write(content)
        print(f"📝 Saved raw data to {debug_file} for inspection")
        
        gbe_lines = []
        for line in content.split('\n'):
            if 'AllWeather' in line:  
                gbe_lines.append(line)
        
        print(f"\n🔍 Found {len(gbe_lines)} GBE lines")
                
        all_accounts_cnav = {}
        all_accounts_positions = {}
        all_accounts_cash = {}
        
        print("\n📊 Processing GBE data...")
        
        for line in gbe_lines:
            parts = [p.strip('"') for p in line.split('","')]
            if len(parts) < 5:
                continue
                
            section_type = parts[0].replace('"', '')
            section_name = parts[1]
            
            if len(parts) > 3:
                current_account = parts[2]
                account_alias = parts[3]
                model = parts[4]
            else:
                continue
            
            if section_type == 'DATA':
                if section_name == 'CNAV':
                    if len(parts) >= 58:
                        try:
                            start_value = float(parts[8] or 0)
                            end_value = float(parts[56] or 0)
                            twr = float(parts[57] or 0) / 100.0 if parts[57] else 0
                            
                            all_accounts_cnav[current_account] = {
                                'start_value': start_value,
                                'end_value': end_value,
                                'twr': twr,
                                'account_id': current_account,
                                'account_alias': account_alias,
                            }
                            
                            primary_marker = " ⭐ PRIMARY" if current_account == self.primary_account else ""
                            print(f"CNAV: {current_account} ({account_alias}) - TWR: {twr:.4f} ({twr*100:.2f}%){primary_marker}")
                            
                        except (ValueError, IndexError) as e:
                            print(f"  Error parsing CNAV for {current_account}: {e}")
                            
                elif section_name == 'MTMP':
                    if len(parts) >= 32:
                        try:
                            # Symbol is at index 17 (trading symbol), not index 7 (ConID)
                            symbol = parts[17] if len(parts) > 17 else parts[7]
                            if symbol and 'Total P/L' not in symbol:
                                close_quantity = float(parts[30] or 0)
                                close_price = float(parts[31] or 0)
                                position_value = close_quantity * close_price
                                
                                if position_value > 0:
                                    if current_account not in all_accounts_positions:
                                        all_accounts_positions[current_account] = {}
                                    
                                    # All non-empty symbols are positions (cash comes from MTM in CNAV)
                                    all_accounts_positions[current_account][symbol] = position_value
                                    print(f"  📈 Position {symbol}: ${position_value:,.2f} (qty: {close_quantity}, price: ${close_price})")
                                    
                        except (ValueError, IndexError) as e:
                            print(f"  ⚠️  Error parsing MTMP: {e}")
                            continue
        
        print(f"\n📋 Data extraction summary:")
        print(f"  Accounts with CNAV: {len(all_accounts_cnav)}")
        print(f"  Accounts with positions: {len(all_accounts_positions)}")
        print(f"  Accounts with cash: {len(all_accounts_cash)}")
        
        return all_accounts_cnav, all_accounts_positions, all_accounts_cash
    
    def aggregate_accounts_data(self, all_accounts_cnav, all_accounts_positions, all_accounts_cash):
        """Focus on primary account"""
        
        print(f"\nProcessing GBE accounts (focusing on primary: {self.primary_account})...")
        
        if self.primary_account in all_accounts_cnav:
            print(f"✅ Using primary account: {self.primary_account}")
            primary_cnav = all_accounts_cnav[self.primary_account]
            primary_positions = all_accounts_positions.get(self.primary_account, {}).copy()
            
            print(f"Primary Account Portfolio: ${primary_cnav['start_value']:,.2f} -> ${primary_cnav['end_value']:,.2f}")
            print(f"Primary Account TWR: {primary_cnav['twr']:.4f} ({primary_cnav['twr']*100:.2f}%)")
            
            if primary_positions:
                total_value = sum(primary_positions.values())
                print(f"\nPrimary Account Positions (Total: ${total_value:,.2f}):")
                for symbol, value in sorted(primary_positions.items(), key=lambda x: x[1], reverse=True):
                    allocation = value / total_value if total_value > 0 else 0
                    print(f"  {symbol}: ${value:,.2f} ({allocation:.1%})")
            
            return primary_cnav, primary_positions
        
        else:
            print(f"⚠️  Primary account {self.primary_account} not found!")
            return None, {}
    
    def calculate_allocations(self, positions):
        total_value = sum(positions.values())
        
        if total_value <= 0:
            return {}
        
        allocations = {}
        for symbol, value in positions.items():
            allocations[symbol] = value / total_value
        
        return allocations
    
    def map_to_canonical_etfs(self, allocations):        
        canonical_allocations = {etf: 0.0 for etf in self.canonical_etfs}
        mapping_used = {}
        unmapped_total = 0.0
        
        print("\n🗺️  Mapping to canonical ETFs:")
        for actual_symbol, allocation in allocations.items():
            canonical_symbol = self.etf_mapping.get(actual_symbol, actual_symbol)
            
            if canonical_symbol in canonical_allocations:
                canonical_allocations[canonical_symbol] += allocation
                mapping_used[actual_symbol] = canonical_symbol
                print(f"  ✅ Mapped {actual_symbol} -> {canonical_symbol}: {allocation:.1%}")
            else:
                unmapped_total += allocation
                print(f"  ❌ UNMAPPED: {actual_symbol} ({allocation:.1%})")
        
        # Validation
        total_mapped = sum(canonical_allocations.values())
        print(f"\n📊 Mapping Summary:")
        print(f"  Total mapped allocation: {total_mapped:.1%}")
        print(f"  Total unmapped allocation: {unmapped_total:.1%}")
        print(f"  Sum: {(total_mapped + unmapped_total):.1%}")
        
        if unmapped_total > 0.01:
            print(f"  ⚠️  WARNING: {unmapped_total:.1%} of portfolio is unmapped!")
        
        return canonical_allocations, mapping_used
    
    def calculate_attribution_and_returns(self, cnav_data, canonical_allocations):
        """Calculate attribution and net returns with fees"""
        
        gross_return = cnav_data.get('twr', 0.0)
        fee_rate = 0.005 
        monthly_fee_rate = fee_rate / 12
        net_return = gross_return - monthly_fee_rate
        
        print(f"\nReturn Calculations:")
        print(f"Gross Return: {gross_return:.4f} ({gross_return*100:.2f}%)")
        print(f"Monthly Fee: {monthly_fee_rate:.4f} ({monthly_fee_rate*100:.2f}%)")
        print(f"Net Return: {net_return:.4f} ({net_return*100:.2f}%)")
        
        attributions = {}
        for etf in self.canonical_etfs:
            allocation_pct = canonical_allocations.get(etf, 0.0)
            attribution = allocation_pct * gross_return
            attributions[f"Attr_{etf}"] = attribution
            if attribution != 0:
                print(f"Attribution {etf}: {allocation_pct:.1%} × {gross_return:.2%} = {attribution:.4f}")
        
        total_attribution = sum(attributions.values())
        print(f"\nAttribution Validation:")
        print(f"Sum of attributions: {total_attribution:.4f}")
        print(f"Gross return: {gross_return:.4f}")
        print(f"Difference: {abs(total_attribution - gross_return):.6f}")
        
        if abs(total_attribution - gross_return) < 0.0001:
            print("✅ Attribution calculations are correct!")
        else:
            print("⚠️  Attribution sum doesn't match total return")
        
        return gross_return, net_return, attributions
    
    def create_gbe_dataframe(self, cnav_data, canonical_allocations, gross_return, net_return, attributions, override_date=None):
        """Create the GBE DataFrame for Excel update
        
        Args:
            override_date: Optional tuple (year, month) to override the date calculation
        """
        
        if override_date:
            year, month = override_date
            last_day_of_month = calendar.monthrange(year, month)[1]
            report_date = datetime(year, month, last_day_of_month)
            print(f"📅 Using override date: {report_date.strftime('%m/%d/%Y')}")
        else:
            today = date.today()
            if today.month == 1:
                last_month = 12
                last_year = today.year - 1
            else:
                last_month = today.month - 1
                last_year = today.year
            
            last_day_of_last_month = calendar.monthrange(last_year, last_month)[1]
            report_date = datetime(last_year, last_month, last_day_of_last_month)
        
        row_data = {
            "Date": report_date,  # Write as datetime object
            "GBE Returns (Gross)": gross_return,
            "GBE Returns (Net)": net_return,
        }
        
        # Add all canonical ETF allocations
        for etf in self.canonical_etfs:
            row_data[f"{etf}_alloc"] = canonical_allocations[etf]
        
        # Add attributions
        row_data.update(attributions)
        
        return pd.DataFrame([row_data])
    
    def update_excel_surgical(self, gbe_df, excel_file="GBEv3.xlsx"):
        """Surgically update Excel file with GBE data"""
        
        print(f"\nUpdating Excel file: {excel_file}")
        
        backup_name = f"GBE_integrated_backup_{date.today().strftime('%Y%m%d')}.xlsx"
        try:
            import shutil
            shutil.copy2(excel_file, backup_name)
            print(f"✅ Created backup: {backup_name}")
        except Exception as e:
            print(f"⚠️  Backup failed: {e}")
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=False)
            ws = wb.active
            
            last_data_row = 0
            for row in range(1, ws.max_row + 1):
                if ws.cell(row=row, column=1).value:
                    last_data_row = row
            
            new_row = last_data_row + 1
            row_data = gbe_df.iloc[0]
            
            print(f"Adding data to row {new_row}")
            print(f"Date: {row_data['Date']}")
            
            # Write date as datetime object with proper Excel date format
            ws.cell(row=new_row, column=1).value = row_data['Date']
            ws.cell(row=new_row, column=1).number_format = 'M/D/YYYY'
            
            ws.cell(row=new_row, column=2).value = row_data['GBE Returns (Gross)']
            ws.cell(row=new_row, column=3).value = row_data['GBE Returns (Net)']
            ws.cell(row=new_row, column=2).number_format = '0.00%'
            ws.cell(row=new_row, column=3).number_format = '0.00%'
            
            # Holdings - columns J-W (10-23) for 14 ETFs - shifted one column to the right
            # PDBC, VWO, IEV, BWX, IAU, VEA, RWX, EWJ, BIV, SPY, TIP, BND, SPTL, VNQ
            visible_etfs = [etf for etf in self.canonical_etfs if etf != 'USFR']
            holdings_cols = list(range(10, 24))  # Columns J-W (10-23)
            holdings_keys = [f"{etf}_alloc" for etf in visible_etfs]
            
            print(f"\n📊 Writing Holdings (columns 10-23):")
            for col, key, etf in zip(holdings_cols, holdings_keys, visible_etfs):
                if key in row_data:
                    ws.cell(row=new_row, column=col).value = row_data[key]
                    ws.cell(row=new_row, column=col).number_format = '0%'
                    print(f"  Col {col} ({etf}): {row_data[key]:.1%}")
            
            # Attribution - columns X-AK (24-37) for 14 ETFs - shifted one column to the right
            attr_cols = list(range(24, 38))  # Columns X-AK (24-37)
            attr_keys = [f"Attr_{etf}" for etf in visible_etfs]
            
            print(f"\n📊 Writing Attribution (columns 24-37):")
            for col, key, etf in zip(attr_cols, attr_keys, visible_etfs):
                if key in row_data:
                    ws.cell(row=new_row, column=col).value = row_data[key]
                    ws.cell(row=new_row, column=col).number_format = '0.00%'
                    print(f"  Col {col} (Attr_{etf}): {row_data[key]:.4f}")
            
            wb.save(excel_file)
            wb.close()
            
            print(f"✅ Successfully updated {excel_file}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error updating Excel file: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def run_monthly_update(self, override_date=None):
        """Run the complete monthly update workflow
        
        Args:
            override_date: Optional tuple (year, month) to override the date calculation
        """
        
        try:
            # Step 1: Extract all GBE data directly from API
            all_accounts_cnav, all_accounts_positions, all_accounts_cash = self.extract_all_weather_data_direct()
            
            if not all_accounts_cnav or not all_accounts_positions:
                print("❌ No GBE data found!")
                return False
            
            # Step 2: Aggregate data across all accounts (including cash)
            aggregated_cnav, aggregated_positions = self.aggregate_accounts_data(all_accounts_cnav, all_accounts_positions, all_accounts_cash)
            
            if not aggregated_cnav:
                print("❌ Failed to aggregate account data!")
                return False
            
            print("\n🔍 Position Breakdown:")
            for symbol, value in sorted(aggregated_positions.items(), key=lambda x: x[1], reverse=True):
                print(f"  {symbol}: ${value:,.2f}")
            
            allocations = self.calculate_allocations(aggregated_positions)
            
            # Validate allocations sum to 100%
            total_allocation = sum(allocations.values())
            print(f"\n✅ Total allocation before mapping: {total_allocation:.2%}")
            if abs(total_allocation - 1.0) > 0.001:
                print(f"  ⚠️  WARNING: Allocations don't sum to 100%! (Sum: {total_allocation:.2%})")
            
            canonical_allocations, mapping_used = self.map_to_canonical_etfs(allocations)
            
            gross_return, net_return, attributions = self.calculate_attribution_and_returns(aggregated_cnav, canonical_allocations)
            
            gbe_df = self.create_gbe_dataframe(aggregated_cnav, canonical_allocations, gross_return, net_return, attributions, override_date)
            
            success = self.update_excel_surgical(gbe_df)
            
            if success:
                print("\n🎉 GBE Monthly Update Completed Successfully!")
                print("Summary:")
                print(f"  - Processed {len(all_accounts_cnav)} GBE accounts")
                print(f"  - Aggregated portfolio value: ${aggregated_cnav['end_value']:,.2f}")
                print(f"  - Monthly return: {gross_return:.2%} (gross), {net_return:.2%} (net)")
                print(f"  - Excel updated with proper formatting")
                return True
            else:
                print("\n❌ GBE Monthly Update Failed!")
                return False
                
        except Exception as e:
            print(f"❌ Error in monthly update: {e}")
            import traceback
            traceback.print_exc()
            return False

def main():
    """Main execution"""
    import sys
    
    updater = GBEMonthlyUpdater()
    
    # Check if specific month is requested
    if len(sys.argv) > 1:
        month_arg = sys.argv[1].lower()
        if month_arg == 'september' or month_arg == 'sep':
            print("\n🗓️  Running for SEPTEMBER 2025 (9/30/25)")
            updater.run_monthly_update(override_date=(2025, 9))
        elif month_arg == 'october' or month_arg == 'oct':
            print("\n🗓️  Running for OCTOBER 2025 (10/31/25)")
            updater.run_monthly_update(override_date=(2025, 10))
        else:
            print(f"Unknown month: {month_arg}")
            print("Usage: python gbe_integrated_updater.py [september|october]")
    else:
        # Default: run for most recent month
        updater.run_monthly_update()

if __name__ == "__main__":
    main()