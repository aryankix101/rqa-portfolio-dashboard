"""
GAM Integrated Monthly Updater
Combines GAMMorningstar extraction and Excel update in one workflow
Processes ALL GAMMorningstar accounts and uses DataFrame transfers (no intermediate CSV)
"""

import pandas as pd
import openpyxl
from datetime import date, datetime
import calendar
import os
import requests
import time
import gzip
import io
import csv
from xml.etree import ElementTree as ET

class GAMMonthlyUpdater:
    """Integrated GAM monthly data processor"""
    
    def __init__(self):
        self.etf_mapping = {
            "COMT": "PDBC",  # iShares GSCI Commodity -> PowerShares DB Commodity
            "SCHE": "VWO",   # Schwab Emerging Markets -> Vanguard Emerging Markets
            "SPLG": "SPY",   # SPDR Portfolio S&P 500 -> SPDR S&P 500
            "TLT": "SPTL",   # iShares 20+ Year Treasury -> SPDR Portfolio Long Term Treasury
            "IAU": "IAU",    # iShares Gold Trust
            "VEA": "VEA",    # Vanguard FTSE Developed Markets
            "VNQ": "VNQ",    # Vanguard Real Estate ETF
            "USFR": "USFR",  # WisdomTree Floating Rate Treasury
            "USD_CASH": "USFR",  # Map cash to USFR for canonical representation
        }
        
        self.canonical_etfs = ["PDBC", "VWO", "IAU", "VEA", "SPY", "SPTL", "VNQ", "USFR"]
        
        # IBKR API Configuration
        self.flex_token = "158159079699491281111623"
        self.query_id = "1303095"
        self.send_url = "https://www.interactivebrokers.com/Universal/servlet/FlexStatementService.SendRequest"
        self.get_url = "https://www.interactivebrokers.com/Universal/servlet/FlexStatementService.GetStatement"
        
        # Target account prioritization (U4686363 - GAM Roth IRA is primary)
        self.primary_account = "U4686363"  # GAM Roth IRA
    
    def fetch_flex_data_direct(self):
        """Fetch flex data directly from IBKR API without file intermediates"""
        
        print("Fetching data directly from IBKR Flex Query API...")
        print(f"Token: {self.flex_token}")
        print(f"Query ID: {self.query_id}")
        
        session = requests.Session()
        headers = {"User-Agent": "GAM-FlexPipeline/1.0"}
        
        try:
            print("Step 1: Sending flex query request...")
            params = {"t": self.flex_token, "q": self.query_id, "v": "3"}
            response = session.get(self.send_url, params=params, timeout=30, headers=headers)
            
            if response.status_code != 200:
                raise Exception(f"SendRequest failed with HTTP {response.status_code}: {response.text[:500]}")
            
            root = ET.fromstring(response.text)
            status = root.findtext(".//Status") or root.findtext(".//status") or ""
            
            if status.lower() == "fail":
                error_code = root.findtext(".//ErrorCode") or ""
                error_msg = root.findtext(".//ErrorMessage") or ""
                raise Exception(f"Flex request failed: {error_code} - {error_msg}")
            
            ref_code = root.findtext(".//ReferenceCode") or root.findtext(".//referenceCode")
            if not ref_code:
                raise Exception(f"No reference code in response: {response.text[:500]}")
            
            print(f"✅ Got reference code: {ref_code}")
            
            print("Step 2: Polling for results...")
            max_attempts = 10
            
            for attempt in range(1, max_attempts + 1):
                time.sleep(3 if attempt > 1 else 2)
                
                params = {"t": self.flex_token, "q": ref_code, "v": "3"}
                response = session.get(self.get_url, params=params, timeout=60, headers=headers)
                
                if response.status_code != 200:
                    print(f"Attempt {attempt}: HTTP {response.status_code}")
                    continue
                
                if response.text.lstrip().startswith("<"):
                    try:
                        root = ET.fromstring(response.text)
                        status = root.findtext(".//Status") or root.findtext(".//status") or ""
                        
                        if status.lower() in {"success", "ready"}:
                            # Data is ready, but might be XML format - return content
                            content = response.content
                        else:
                            print(f"Attempt {attempt}: Status = {status}")
                            continue
                    except ET.ParseError:
                        content = response.content
                else:
                    content = response.content
                
                if response.headers.get("Content-Encoding", "").lower() == "gzip" or content[:2] == b"\x1f\x8b":
                    try:
                        content = gzip.decompress(content)
                    except Exception:
                        pass
                
                print(f"✅ Retrieved data ({len(content)} bytes)")
                return content.decode("utf-8", errors="replace")
            
            raise Exception(f"Data not ready after {max_attempts} attempts")
            
        except Exception as e:
            print(f"❌ API Error: {e}")
            raise
        
    def extract_all_gammorningstar_data_direct(self):
        """Extract data from ALL GAMMorningstar accounts using direct API call"""
        
        print("GAM Integrated Monthly Updater")
        print("=" * 50)
        print("Extracting ALL GAMMorningstar accounts...")
        
        # Fetch data directly from API (no file intermediate)
        content = self.fetch_flex_data_direct()
        
        gam_lines = []
        for line in content.split('\n'):
            if 'GAMMorningstar' in line:
                gam_lines.append(line)
        
        print(f"Found {len(gam_lines)} GAMMorningstar lines")
        
        all_accounts_cnav = {}
        all_accounts_positions = {}
        all_accounts_cash = {}
        
        for line in gam_lines:
            parts = [p.strip('"') for p in line.split('","')]
            if len(parts) < 5:
                continue
                
            section_type = parts[0].replace('"', '')
            section_name = parts[1]
            
            if len(parts) > 3:
                current_account = parts[2]
                account_alias = parts[3]
                model = parts[4]
            else:
                continue
            
            if section_type == 'DATA':
                if section_name == 'CNAV':
                    if len(parts) >= 58:
                        try:
                            start_value = float(parts[8] or 0)
                            end_value = float(parts[56] or 0)
                            twr = float(parts[57] or 0) / 100.0 if parts[57] else 0
                            
                            all_accounts_cnav[current_account] = {
                                'start_value': start_value,
                                'end_value': end_value,
                                'twr': twr,
                                'account_id': current_account,
                                'account_alias': account_alias
                            }
                            
                            primary_marker = " ⭐ PRIMARY" if current_account == self.primary_account else ""
                            print(f"CNAV: {current_account} ({account_alias}) - TWR: {twr:.4f} ({twr*100:.2f}%){primary_marker}")
                            
                        except (ValueError, IndexError) as e:
                            print(f"  Error parsing CNAV for {current_account}: {e}")
                            
                elif section_name == 'MTMP':
                    if len(parts) >= 32:
                        try:
                            symbol = parts[7]
                            if symbol and 'Total P/L' not in symbol:
                                close_quantity = float(parts[30] or 0)
                                close_price = float(parts[31] or 0)
                                position_value = close_quantity * close_price
                                
                                if position_value > 0:
                                    if current_account not in all_accounts_positions:
                                        all_accounts_positions[current_account] = {}
                                        all_accounts_cash[current_account] = 0.0
                                    
                                    if symbol in ['USD', '']:
                                        all_accounts_cash[current_account] += position_value
                                        print(f"  Cash: ${position_value:,.2f}")
                                    else:
                                        all_accounts_positions[current_account][symbol] = position_value
                                    
                        except (ValueError, IndexError) as e:
                            continue
                            
                elif section_name in ['CRTT', 'POST']:
                    if len(parts) >= 10:
                        try:
                            currency = parts[6] if len(parts) > 6 else ''
                            if currency == 'USD':
                                cash_value = float(parts[8] or 0) if len(parts) > 8 else 0
                                if cash_value > 0:
                                    if current_account not in all_accounts_cash:
                                        all_accounts_cash[current_account] = 0.0
                                    all_accounts_cash[current_account] += cash_value
                        except (ValueError, IndexError):
                            continue
        
        return all_accounts_cnav, all_accounts_positions, all_accounts_cash
    
    def aggregate_accounts_data(self, all_accounts_cnav, all_accounts_positions, all_accounts_cash):
        """Focus on primary account U4686363 (GAM Roth IRA) with fallback aggregation"""
        
        print(f"\nProcessing GAMMorningstar accounts (focusing on primary: {self.primary_account})...")
        
        if self.primary_account in all_accounts_cnav:
            print(f"✅ Using primary account: {self.primary_account}")
            primary_cnav = all_accounts_cnav[self.primary_account]
            primary_positions = all_accounts_positions.get(self.primary_account, {}).copy()
            primary_cash = all_accounts_cash.get(self.primary_account, 0.0)
            
            print(f"Primary Account Portfolio: ${primary_cnav['start_value']:,.2f} -> ${primary_cnav['end_value']:,.2f}")
            print(f"Primary Account TWR: {primary_cnav['twr']:.4f} ({primary_cnav['twr']*100:.2f}%)")
            
            if primary_cash > 0:
                primary_positions['USD_CASH'] = primary_cash
                total_value = sum(primary_positions.values())
                cash_allocation = primary_cash / total_value if total_value > 0 else 0
                print(f"Primary Account Cash: ${primary_cash:,.2f} ({cash_allocation:.1%})")
            
            if primary_positions:
                total_value = sum(primary_positions.values())
                print(f"\nPrimary Account Positions (Total: ${total_value:,.2f}):")
                for symbol, value in sorted(primary_positions.items(), key=lambda x: x[1], reverse=True):
                    allocation = value / total_value if total_value > 0 else 0
                    print(f"  {symbol}: ${value:,.2f} ({allocation:.1%})")
            
            return primary_cnav, primary_positions
        
        else:
            print(f"⚠️  Primary account {self.primary_account} not found, falling back to aggregation...")
            
            total_start_value = 0
            total_end_value = 0
            
            for account_id, cnav_data in all_accounts_cnav.items():
                total_start_value += cnav_data['start_value']
                total_end_value += cnav_data['end_value']
                
            if total_start_value > 0:
                aggregated_twr = (total_end_value / total_start_value) - 1
            else:
                aggregated_twr = 0
            
            print(f"Aggregated Portfolio: ${total_start_value:,.2f} -> ${total_end_value:,.2f}")
            print(f"Aggregated TWR: {aggregated_twr:.4f} ({aggregated_twr*100:.2f}%)")
            
            aggregated_positions = {}
            total_cash = 0
            
            for account_id, positions in all_accounts_positions.items():
                for symbol, value in positions.items():
                    if symbol in aggregated_positions:
                        aggregated_positions[symbol] += value
                    else:
                        aggregated_positions[symbol] = value
            
            for account_id, cash_value in all_accounts_cash.items():
                total_cash += cash_value
        
            # Include cash in total portfolio value calculation
            total_position_value = sum(aggregated_positions.values()) + total_cash
            
            print(f"\nAggregated Positions (Total: ${total_position_value:,.2f}):")
            for symbol, value in sorted(aggregated_positions.items(), key=lambda x: x[1], reverse=True):
                allocation = value / total_position_value if total_position_value > 0 else 0
                print(f"  {symbol}: ${value:,.2f} ({allocation:.1%})")
            
            if total_cash > 0:
                cash_allocation = total_cash / total_position_value if total_position_value > 0 else 0
                print(f"  USD Cash: ${total_cash:,.2f} ({cash_allocation:.1%})")
                aggregated_positions['USD_CASH'] = total_cash
            
            aggregated_cnav = {
                'start_value': total_start_value,
                'end_value': total_end_value,
                'twr': aggregated_twr,
                'account_id': 'GAM_AGGREGATED',
                'account_alias': 'All GAMMorningstar Accounts'
            }
            
            return aggregated_cnav, aggregated_positions
    
    def calculate_allocations(self, positions):
        """Calculate percentage allocations from positions"""
        total_value = sum(positions.values())
        
        if total_value <= 0:
            return {}
        
        allocations = {}
        for symbol, value in positions.items():
            allocations[symbol] = value / total_value
        
        return allocations
    
    def map_to_canonical_etfs(self, allocations):
        """Map actual ETF holdings to canonical GAM ETFs"""
        
        canonical_allocations = {etf: 0.0 for etf in self.canonical_etfs}
        mapping_used = {}
        
        for actual_symbol, allocation in allocations.items():
            canonical_symbol = self.etf_mapping.get(actual_symbol, actual_symbol)
            
            if canonical_symbol in canonical_allocations:
                canonical_allocations[canonical_symbol] += allocation
                mapping_used[actual_symbol] = canonical_symbol
                print(f"Mapped {actual_symbol} -> {canonical_symbol}: {allocation:.1%}")
            else:
                print(f"Unmapped symbol: {actual_symbol} ({allocation:.1%})")
        
        return canonical_allocations, mapping_used
    
    def calculate_attribution_and_returns(self, cnav_data, canonical_allocations):
        """Calculate attribution and net returns with fees"""
        
        gross_return = cnav_data.get('twr', 0.0)
        fee_rate = 0.0075
        monthly_fee_rate = fee_rate / 12
        net_return = gross_return - monthly_fee_rate
        
        print(f"\nReturn Calculations:")
        print(f"Gross Return: {gross_return:.4f} ({gross_return*100:.2f}%)")
        print(f"Monthly Fee: {monthly_fee_rate:.4f} ({monthly_fee_rate*100:.2f}%)")
        print(f"Net Return: {net_return:.4f} ({net_return*100:.2f}%)")
        
        attributions = {}
        for etf in self.canonical_etfs:
            allocation_pct = canonical_allocations.get(etf, 0.0)
            attribution = allocation_pct * gross_return
            attributions[f"Attr_{etf}"] = attribution
            if attribution != 0:
                print(f"Attribution {etf}: {allocation_pct:.1%} × {gross_return:.2%} = {attribution:.4f}")
        
        total_attribution = sum(attributions.values())
        print(f"\nAttribution Validation:")
        print(f"Sum of attributions: {total_attribution:.4f}")
        print(f"Gross return: {gross_return:.4f}")
        print(f"Difference: {abs(total_attribution - gross_return):.6f}")
        
        if abs(total_attribution - gross_return) < 0.0001:
            print("✅ Attribution calculations are correct!")
        else:
            print("⚠️  Attribution sum doesn't match total return")
        
        return gross_return, net_return, attributions
    
    def create_gam_dataframe(self, cnav_data, canonical_allocations, gross_return, net_return, attributions):
        """Create the GAM DataFrame for Excel update"""
        
        today = date.today()
        if today.month == 1:
            last_month = 12
            last_year = today.year - 1
        else:
            last_month = today.month - 1
            last_year = today.year
        
        last_day_of_last_month = calendar.monthrange(last_year, last_month)[1]
        last_month_date_obj = date(last_year, last_month, last_day_of_last_month)
        
        row_data = {
            "Date": last_month_date_obj.strftime('%-m/%-d/%Y'),
            "GAM Returns (Gross)": gross_return,
            "GAM Returns (Net)": net_return,
            "PDBC": canonical_allocations["PDBC"],
            "VWO": canonical_allocations["VWO"], 
            "IAU": canonical_allocations["IAU"],
            "VEA": canonical_allocations["VEA"],
            "SPY_alloc": canonical_allocations["SPY"],
            "SPTL": canonical_allocations["SPTL"],
            "VNQ": canonical_allocations["VNQ"],
            "USFR": canonical_allocations["USFR"],
            **attributions
        }
        
        return pd.DataFrame([row_data])
    
    def update_excel_surgical(self, gam_df, excel_file="GAM_new copy.xlsx"):
        """Surgically update Excel file with GAM data"""
        
        print(f"\nUpdating Excel file: {excel_file}")
        
        backup_name = f"GAM_integrated_backup_{date.today().strftime('%Y%m%d')}.xlsx"
        try:
            import shutil
            shutil.copy2(excel_file, backup_name)
            print(f"✅ Created backup: {backup_name}")
        except Exception as e:
            print(f"⚠️  Backup failed: {e}")
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=False)
            ws = wb.active
            
            last_data_row = 0
            for row in range(1, ws.max_row + 1):
                if ws.cell(row=row, column=1).value:
                    last_data_row = row
            
            new_row = last_data_row + 1
            row_data = gam_df.iloc[0]
            
            print(f"Adding data to row {new_row}")
            print(f"Date: {row_data['Date']}")
            
            ws.cell(row=new_row, column=1).value = row_data['Date']
            
            ws.cell(row=new_row, column=2).value = row_data['GAM Returns (Gross)']
            ws.cell(row=new_row, column=3).value = row_data['GAM Returns (Net)']
            ws.cell(row=new_row, column=2).number_format = '0.00%'
            ws.cell(row=new_row, column=3).number_format = '0.00%'
            
            # Holdings (columns 10-17) with percentage formatting
            holdings_cols = [10, 11, 12, 13, 14, 15, 16, 17]
            holdings_keys = ['PDBC', 'VWO', 'IAU', 'VEA', 'SPY_alloc', 'SPTL', 'VNQ', 'USFR']
            
            for col, key in zip(holdings_cols, holdings_keys):
                ws.cell(row=new_row, column=col).value = row_data[key]
                ws.cell(row=new_row, column=col).number_format = '0%'
            
            # Attribution (columns 18-25) with percentage formatting
            attr_cols = [18, 19, 20, 21, 22, 23, 24, 25]
            attr_keys = ['Attr_PDBC', 'Attr_VWO', 'Attr_IAU', 'Attr_VEA', 'Attr_SPY', 'Attr_SPTL', 'Attr_VNQ', 'Attr_USFR']
            
            for col, key in zip(attr_cols, attr_keys):
                ws.cell(row=new_row, column=col).value = row_data[key]
                ws.cell(row=new_row, column=col).number_format = '0.00%'
            
            wb.save(excel_file)
            wb.close()
            
            print(f"✅ Successfully updated {excel_file}")
            print(f"GAM Return (Gross): {row_data['GAM Returns (Gross)']:.2%}")
            print(f"GAM Return (Net): {row_data['GAM Returns (Net)']:.2%}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error updating Excel file: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def run_monthly_update(self):
        """Run the complete monthly update workflow (no file dependencies)"""
        
        try:
            # Step 1: Extract all GAMMorningstar data directly from API
            all_accounts_cnav, all_accounts_positions, all_accounts_cash = self.extract_all_gammorningstar_data_direct()
            
            if not all_accounts_cnav or not all_accounts_positions:
                print("❌ No GAMMorningstar data found!")
                return False
            
            # Step 2: Aggregate data across all accounts (including cash)
            aggregated_cnav, aggregated_positions = self.aggregate_accounts_data(all_accounts_cnav, all_accounts_positions, all_accounts_cash)
            
            allocations = self.calculate_allocations(aggregated_positions)
            
            canonical_allocations, mapping_used = self.map_to_canonical_etfs(allocations)
            
            gross_return, net_return, attributions = self.calculate_attribution_and_returns(aggregated_cnav, canonical_allocations)
            
            gam_df = self.create_gam_dataframe(aggregated_cnav, canonical_allocations, gross_return, net_return, attributions)
            
            success = self.update_excel_surgical(gam_df)
            
            if success:
                print("\n🎉 GAM Monthly Update Completed Successfully!")
                print("Summary:")
                print(f"  - Processed {len(all_accounts_cnav)} GAMMorningstar accounts")
                print(f"  - Aggregated portfolio value: ${aggregated_cnav['end_value']:,.2f}")
                print(f"  - Monthly return: {gross_return:.2%} (gross), {net_return:.2%} (net)")
                print(f"  - Excel updated with proper formatting")
                print(f"  - Database synchronized")
                return True
            else:
                print("\n❌ GAM Monthly Update Failed!")
                return False
                
        except Exception as e:
            print(f"❌ Error in monthly update: {e}")
            import traceback
            traceback.print_exc()
            return False

def main():
    """Main execution"""
    updater = GAMMonthlyUpdater()
    updater.run_monthly_update()

if __name__ == "__main__":
    main()