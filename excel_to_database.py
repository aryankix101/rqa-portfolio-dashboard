import pandas as pd
import sqlite3
import numpy as np
from datetime import datetime
import openpyxl

# Database setup
def create_database():
    conn = sqlite3.connect('portfolio_data.db')
    cursor = conn.cursor()
    
    # Create monthly_returns table - matches your Monthly Returns chart
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS monthly_returns (
        date TEXT PRIMARY KEY,
        ga_returns_gross REAL,
        ga_returns_net REAL,
        acwi REAL,
        agg REAL,
        spy REAL,
        portfolio_50_50 REAL,
        portfolio_60_40 REAL,
        portfolio_70_30 REAL
    )
    ''')
    
    # Create ga_allocations table - for the Trailing 12 Month Allocations chart
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS ga_allocations (
        date TEXT,
        asset_symbol TEXT,
        asset_name TEXT,
        allocation_percentage REAL,
        PRIMARY KEY (date, asset_symbol)
    )
    ''')
    
    # Create ga_attribution table - for the Portfolio Attribution chart
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS ga_attribution (
        date TEXT,
        asset_symbol TEXT,
        asset_name TEXT,
        attribution_value REAL,
        PRIMARY KEY (date, asset_symbol)
    )
    ''')
    
    # Create benchmark_performance table - for the Benchmark Performance table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS benchmark_performance (
        portfolio TEXT PRIMARY KEY,
        ytd REAL,
        one_year REAL,
        five_year REAL,
        since_inception REAL,
        standard_deviation REAL,
        sharpe_ratio REAL,
        beta_to_sp500 REAL
    )
    ''')
    
    # ============================================
    # GBE Strategy Tables (mirrors GA structure)
    # ============================================
    
    # Create gbe_monthly_returns table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS gbe_monthly_returns (
        date TEXT PRIMARY KEY,
        gbe_returns_gross REAL,
        gbe_returns_net REAL,
        acwi REAL,
        agg REAL,
        spy REAL,
        portfolio_50_50 REAL,
        portfolio_60_40 REAL,
        portfolio_70_30 REAL
    )
    ''')
    
    # Create gbe_allocations table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS gbe_allocations (
        date TEXT,
        asset_symbol TEXT,
        asset_name TEXT,
        allocation_percentage REAL,
        PRIMARY KEY (date, asset_symbol)
    )
    ''')
    
    # Create gbe_attribution table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS gbe_attribution (
        date TEXT,
        asset_symbol TEXT,
        asset_name TEXT,
        attribution_value REAL,
        PRIMARY KEY (date, asset_symbol)
    )
    ''')
    
    conn.commit()
    return conn

def read_ga_excel_data():
    wb = openpyxl.load_workbook('GA_new copy.xlsx')
    ws = wb.active
    
    asset_mapping = {
        'PDBC': 'Commodities',
        'VWO': 'EM_Stocks', 
        'IAU': 'Gold',
        'VEA': 'Intl_Dev_Stocks',
        'SPY': 'SP500',
        'SPTL': 'US_LT_Treas',
        'VNQ': 'US_REITs',
        'USFR': 'Cash'
    }
    
    monthly_returns_data = []
    ga_allocations_data = []
    ga_attribution_data = []
    # Start from row 3 (row 2 has headers, row 1 has section labels)
    for row in range(3, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        elif isinstance(date_val, str):
            try:
                dt = datetime.strptime(date_val, '%m/%d/%y')
                date_str = dt.strftime('%Y-%m-%d')
            except Exception:
                continue
        else:
            continue
        
        # Convert numeric strings to floats, handle None values
        def safe_float_convert(value):
            if value is None or value == '' or (isinstance(value, str) and value.strip() == ''):
                return None
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
        
        # Extract monthly returns data (columns 2-9)
        ga_returns_gross = safe_float_convert(ws.cell(row=row, column=2).value)
        # Calculate GA Returns (Net) by subtracting the management fee (0.75% annually = 0.0625% monthly)
        monthly_fee = 0.0075 / 12
        ga_returns_net = ga_returns_gross - monthly_fee if ga_returns_gross is not None else None
        
        # Convert percentage values to decimal (divide by 100) for consistency
        acwi = safe_float_convert(ws.cell(row=row, column=4).value)
        
        agg = safe_float_convert(ws.cell(row=row, column=5).value)
        
        spy = safe_float_convert(ws.cell(row=row, column=6).value)
        
        port_50_50 = safe_float_convert(ws.cell(row=row, column=7).value)
        
        port_60_40 = safe_float_convert(ws.cell(row=row, column=8).value)
        
        port_70_30 = safe_float_convert(ws.cell(row=row, column=9).value)
        
        monthly_returns_data.append({
            'date': date_str,
            'ga_returns_gross': ga_returns_gross,
            'ga_returns_net': ga_returns_net,
            'acwi': acwi,
            'agg': agg,
            'spy': spy,
            'portfolio_50_50': port_50_50,
            'portfolio_60_40': port_60_40,
            'portfolio_70_30': port_70_30
        })
        
        # Extract GA allocation data (columns 10-17)
        allocation_columns = [10, 11, 12, 13, 14, 15, 16, 17]
        for col in allocation_columns:
            asset_symbol = ws.cell(row=2, column=col).value
            allocation_value = safe_float_convert(ws.cell(row=row, column=col).value)
            
            if asset_symbol and allocation_value is not None:
                asset_name = asset_mapping.get(asset_symbol, asset_symbol)
                ga_allocations_data.append({
                    'date': date_str,
                    'asset_symbol': asset_symbol,
                    'asset_name': asset_name,
                    'allocation_percentage': allocation_value
                })
        
        # Extract GA attribution data (columns 18-25)
        attribution_columns = [18, 19, 20, 21, 22, 23, 24, 25]
        for col in attribution_columns:
            asset_symbol = ws.cell(row=2, column=col).value
            attribution_value = safe_float_convert(ws.cell(row=row, column=col).value)
            
            if asset_symbol and attribution_value is not None:
                asset_name = asset_mapping.get(asset_symbol, asset_symbol)
                ga_attribution_data.append({
                    'date': date_str,
                    'asset_symbol': asset_symbol,
                    'asset_name': asset_name,
                    'attribution_value': attribution_value
                })
    
    return (pd.DataFrame(monthly_returns_data), 
            pd.DataFrame(ga_allocations_data), 
            pd.DataFrame(ga_attribution_data))

def read_gbe_excel_data():
    """Read GBE.xlsx and extract monthly returns, allocations, and attribution data"""
    # Load with data_only=True to read calculated formula values instead of formulas
    wb = openpyxl.load_workbook('GBE.xlsx', data_only=True)
    ws = wb.active
    
    # GBE asset mapping - 14 ETFs
    asset_mapping = {
        'PDBC': 'Commodities',
        'VWO': 'EM_Stocks',
        'IEV': 'Europe_Stocks',
        'BWX': 'Intl_Bonds',
        'IAU': 'Gold',
        'VEA': 'Intl_Dev_Stocks',
        'RWX': 'Intl_REITs',
        'EWJ': 'Japan_Stocks',
        'BIV': 'Intermediate_Bonds',
        'SPY': 'SP500',
        'TIP': 'TIPS',
        'BND': 'Agg_Bonds',
        'SPTL': 'US_LT_Treas',
        'VNQ': 'US_REITs'
    }
    
    monthly_returns_data = []
    gbe_allocations_data = []
    gbe_attribution_data = []
    
    # Start from row 3 (row 2 has headers, row 1 has section labels)
    for row in range(3, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        elif isinstance(date_val, str):
            try:
                dt = datetime.strptime(date_val, '%m/%d/%y')
                date_str = dt.strftime('%Y-%m-%d')
            except Exception:
                try:
                    dt = datetime.strptime(date_val, '%m/%d/%Y')
                    date_str = dt.strftime('%Y-%m-%d')
                except Exception:
                    continue
        else:
            continue
        
        # Convert numeric strings to floats, handle None values
        def safe_float_convert(value):
            if value is None or value == '' or (isinstance(value, str) and value.strip() == ''):
                return None
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
        
        # Extract monthly returns data (columns 2-3 for GBE, then benchmark columns)
        gbe_returns_gross = safe_float_convert(ws.cell(row=row, column=2).value)
        gbe_returns_net = safe_float_convert(ws.cell(row=row, column=3).value)
        
        # Benchmarks (assuming similar structure to GA - adjust if different)
        acwi = safe_float_convert(ws.cell(row=row, column=4).value)
        agg = safe_float_convert(ws.cell(row=row, column=5).value)
        spy = safe_float_convert(ws.cell(row=row, column=6).value)
        port_50_50 = safe_float_convert(ws.cell(row=row, column=7).value)
        port_60_40 = safe_float_convert(ws.cell(row=row, column=8).value)
        port_70_30 = safe_float_convert(ws.cell(row=row, column=9).value)
        
        monthly_returns_data.append({
            'date': date_str,
            'gbe_returns_gross': gbe_returns_gross,
            'gbe_returns_net': gbe_returns_net,
            'acwi': acwi,
            'agg': agg,
            'spy': spy,
            'portfolio_50_50': port_50_50,
            'portfolio_60_40': port_60_40,
            'portfolio_70_30': port_70_30
        })
        
        # Extract GBE allocation data (columns 10-23 for 14 ETFs)
        allocation_columns = list(range(10, 24))  # Columns J-W (10-23)
        for col in allocation_columns:
            asset_symbol = ws.cell(row=2, column=col).value
            allocation_value = safe_float_convert(ws.cell(row=row, column=col).value)
            
            if asset_symbol and allocation_value is not None:
                asset_name = asset_mapping.get(asset_symbol, asset_symbol)
                gbe_allocations_data.append({
                    'date': date_str,
                    'asset_symbol': asset_symbol,
                    'asset_name': asset_name,
                    'allocation_percentage': allocation_value
                })
        
        # Extract GBE attribution data (columns 24-37 for 14 ETFs)
        attribution_columns = list(range(24, 38))  # Columns X-AK (24-37)
        for col in attribution_columns:
            asset_symbol = ws.cell(row=2, column=col).value
            attribution_value = safe_float_convert(ws.cell(row=row, column=col).value)
            
            if asset_symbol and attribution_value is not None:
                asset_name = asset_mapping.get(asset_symbol, asset_symbol)
                gbe_attribution_data.append({
                    'date': date_str,
                    'asset_symbol': asset_symbol,
                    'asset_name': asset_name,
                    'attribution_value': attribution_value
                })
    
    return (pd.DataFrame(monthly_returns_data), 
            pd.DataFrame(gbe_allocations_data), 
            pd.DataFrame(gbe_attribution_data))

def annualized_return(returns, periods):
    if len(returns) < periods:
        return None
    period_returns = returns.tail(periods)
    cumulative_return = (1 + period_returns).prod()
    years = periods / 12
    return cumulative_return ** (1/years) - 1

def annualized_std(returns):
    return returns.std() * np.sqrt(12)

def sharpe_ratio(returns):
    if len(returns) == 0:
        return None
    
    cumulative_return = (1 + returns).prod()
    years = len(returns) / 12
    annualized_return = cumulative_return ** (1/years) - 1 if years > 0 else 0
    
    annualized_volatility = returns.std() * np.sqrt(12)
    
    if annualized_volatility == 0:
        return None
    
    return annualized_return / annualized_volatility

def calculate_beta(portfolio_returns, spy_returns):
    if len(portfolio_returns) == 0 or len(spy_returns) == 0:
        return None
    covariance = np.cov(portfolio_returns, spy_returns)[0][1]
    spy_variance = np.var(spy_returns)
    if spy_variance == 0:
        return None
    return covariance / spy_variance

def ytd_return(returns, dates):
    current_year = datetime.now().year
    ytd_data = returns[dates.dt.year == current_year]
    if len(ytd_data) == 0:
        return None
    return (1 + ytd_data).prod() - 1

def store_all_data(conn, monthly_returns_df, ga_allocations_df, ga_attribution_df):
    monthly_returns_df.to_sql('monthly_returns', conn, if_exists='replace', index=False)
    ga_allocations_df.to_sql('ga_allocations', conn, if_exists='replace', index=False)
    ga_attribution_df.to_sql('ga_attribution', conn, if_exists='replace', index=False)

def store_gbe_data(conn, gbe_monthly_returns_df, gbe_allocations_df, gbe_attribution_df):
    """Store GBE data in the database"""
    gbe_monthly_returns_df.to_sql('gbe_monthly_returns', conn, if_exists='replace', index=False)
    gbe_allocations_df.to_sql('gbe_allocations', conn, if_exists='replace', index=False)
    gbe_attribution_df.to_sql('gbe_attribution', conn, if_exists='replace', index=False)


def calculate_benchmark_performance(conn, monthly_returns_df):
    df = monthly_returns_df.copy()
    df['date'] = pd.to_datetime(df['date'])
    df = df.sort_values('date').dropna()
    
    portfolios = {
        'GA': 'ga_returns_net',
        '60/40': 'portfolio_60_40', 
        '70/30': 'portfolio_70_30'
    }
    
    performance_data = []
    
    for portfolio_name, column_name in portfolios.items():
        returns = df[column_name].dropna()
        if len(returns) == 0:
            continue
        
        dates = df.loc[returns.index, 'date']
        spy_returns = df.loc[returns.index, 'spy'].dropna()
        
        ytd = ytd_return(returns, dates)
        one_year = annualized_return(returns, 12)
        five_year = annualized_return(returns, 60)
        since_inception = annualized_return(returns, len(returns))
        standard_deviation = annualized_std(returns)
        sharpe_ratio_val = sharpe_ratio(returns)
        beta = calculate_beta(returns, spy_returns) if column_name != 'spy' else 1.0
        
        performance_data.append({
            'portfolio': portfolio_name,
            'ytd': ytd,
            'one_year': one_year,
            'five_year': five_year,
            'since_inception': since_inception,
            'standard_deviation': standard_deviation,
            'sharpe_ratio': sharpe_ratio_val,
            'beta_to_sp500': beta
        })
    
    performance_df = pd.DataFrame(performance_data)
    performance_df.to_sql('benchmark_performance', conn, if_exists='replace', index=False)

def get_trailing_12m_allocations_chart_data(conn, months=12):
    query = '''
    SELECT date, asset_name, allocation_percentage
    FROM ga_allocations 
    ORDER BY date DESC, asset_name
    LIMIT ?
    '''
    
    cursor = conn.cursor()
    cursor.execute(query, (months * 8,))
    results = cursor.fetchall()
    
    chart_data = {}
    for date, asset_name, percentage in results:
        if date not in chart_data:
            chart_data[date] = {}
        chart_data[date][asset_name] = percentage
    
    return chart_data

def get_attribution_chart_data(conn, months=12):
    query = '''
    SELECT date, asset_name, attribution_value
    FROM ga_attribution 
    ORDER BY date DESC, asset_name
    LIMIT ?
    '''
    
    cursor = conn.cursor()
    cursor.execute(query, (months * 8,))
    results = cursor.fetchall()
    
    chart_data = {}
    for date, asset_name, attribution_value in results:
        if date not in chart_data:
            chart_data[date] = {}
        chart_data[date][asset_name] = attribution_value
    
    return chart_data

def main():
    print("Creating database...")
    conn = create_database()
    
    print("Reading GA Excel data...")
    monthly_returns_df, ga_allocations_df, ga_attribution_df = read_ga_excel_data()
    
    print("Storing GA data...")
    store_all_data(conn, monthly_returns_df, ga_allocations_df, ga_attribution_df)
    
    print("Calculating GA benchmark performance...")
    calculate_benchmark_performance(conn, monthly_returns_df)
    
    print("\nReading GBE Excel data...")
    try:
        gbe_monthly_returns_df, gbe_allocations_df, gbe_attribution_df = read_gbe_excel_data()
        
        print("Storing GBE data...")
        store_gbe_data(conn, gbe_monthly_returns_df, gbe_allocations_df, gbe_attribution_df)
        
        print("GBE data migration complete!")
    except FileNotFoundError:
        print("⚠️  GBE.xlsx not found - skipping GBE data migration")
    except Exception as e:
        print(f"⚠️  Error processing GBE data: {e}")
    
    print("\n✅ Database creation complete!")
    
    conn.close()

if __name__ == "__main__":
    main()